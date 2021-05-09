from machine import LFP
import openpyxl as op
import datetime

CleanRoom = []
UT = []
RT = []
IT = []
VT = []
wb = None


def ReadData():
    global wb
    c = 0  # counter
    wb = op.load_workbook("Status.xlsx")
    ws = wb["Status"]
    wd = wb["StatusByDate"]
    wsn = wb["SN"]
    for i in ws["A"]:
        Status = []
        StatusByDate = []
        SN = []
        if c == 0:
            c = c+1
            continue
        else:
            mac = LFP(i.value)
            for j in ws[str(i.value+1)]:
                if j:
                    Status.append(j.value)
            for k in wd[str(i.value+1)]:
                if k:
                    StatusByDate.append(k.value)
            mac.LoadStatus(Status, StatusByDate)
            for m in wsn[str(i.value+1)]:
                if m:
                    SN.append(m.value)
            mac.LoadSN(SN)
            if wd[str(i.value+1)][5].value:
                mac.LoadDone(wd[str(i.value+1)][5].value)
            # Next check the status e.g uninstalled, installed, and Resistance
            CleanRoom.append(mac)


def UninstallToday():
    global UT
    UT = input("Today Uninstall: ").split()
    UT.sort()


def ResistanceToday():
    global RT
    RT = input("Today Resistance: ").split()
    RT.sort()


def InstallToday():
    global IT
    IT = input("Today Install: ").split()
    IT.sort()


def VerifyToday():
    global VT
    VT = input("VerifyToday: ").split()
    VT.sort()


def UpdateToday():
    print("RELOAD")
    for L in CleanRoom:
        if (str(L.No)in UT):
            L.Uninstall()
        if (str(L.No) in RT):
            L.ReplaceResistance()
        if (str(L.No) in IT):
            L.Install()
        if (str(L.No) in VT):
            L.Verify()


def CheckToday(method):
    if method == "UT":
        print("Today UNINSTALL {}.".format(len(UT)))
        for i in UT:
            print("TP{0:03}".format(int(i)))
    elif method == "RT":
        print("Today replace {} RESISTANCE.".format(len(RT)))
        for i in RT:
            print("TP{0:03}".format(int(i)))
    elif method == "IT":
        print("Today INSTALL {}.".format(len(IT)))
        for i in IT:
            print("TP{0:03}".format(int(i)))
    elif method == "VT":
        print("Today VERIFY {lenth}.".format(lenth=len(VT)))
        for i in VT:
            print("TP{0:03}".format(int(i)))


def KeyInSN():
    id = int(input("Input the machine ID: "))
    for i in CleanRoom:
        if id == i.No:
            SN = input("Key In the S/N Number: ").split()
            print(SN)
            i.LoadSN(SN)
            break


def WriteStatus():
    for row in wb["Status"].iter_rows(min_row=2, max_col=6, max_row=516):
        where = (row[0].value-1)
        if CleanRoom[where].Uninstalled:
            if not row[1].value:
                row[1].value = 'v'
            else:
                print("WARNING: Uninstall Data COMFLICT at TP{0:03}".format(row[0].value))
        if CleanRoom[where].Resistance:
            if not row[2].value:
                row[2].value = 'v'
            else:
                print("WARNING: Resistance Data COMFLICT at TP{0:03}".format(row[0].value))
        if CleanRoom[where].Installed:
            if not row[3].value:
                row[3].value = 'v'
            else:
                print("WARNING: Install Data COMFLICT at TP{0:03}".format(row[0].value))
        if CleanRoom[where].Verified:
            if not row[4].value:
                row[4].value = 'v'
            else:
                print("WARNING: Verify Data COMFLICT at TP{0:03}".format(row[0].value))
        if CleanRoom[where].Done():
            if not row[5].value:
                row[5].value = 'v'
            else:
                print("WARNING: Finish Data COMFLICT at TP{0:03}".format(row[0].value))


def WriteStatusByDate():
    for row in wb["StatusByDate"].iter_rows(min_row=2, max_col=6, max_row=516):
        where = row[0].value-1
        if CleanRoom[where].Uninstalled:
            if not row[1].value:
                row[1].value = datetime.date.today()
            else:
                print("WARNING: Uninstall Data COMFLICT at TP{0:03}".format(row[0].value))
        if CleanRoom[where].Resistance:
            if not row[2].value:
                row[2].value = datetime.date.today()
            else:
                print("WARNING: Resistance Data COMFLICT at TP{0:03}".format(row[0].value))
        if CleanRoom[where].Installed:
            if not row[3].value:
                row[3].value = datetime.date.today()
            else:
                print("WARNING: Install Data COMFLICT at TP{0:03}".format(row[0].value))
        if CleanRoom[where].Verified:
            if not row[4].value:
                row[4].value = datetime.date.today()
            else:
                print("WARNING: Verify Data COMFLICT at TP{0:03}".format(row[0].value))
        if CleanRoom[where].Done():
            if not row[5].value:
                row[5].value = datetime.date.today()
            else:
                print("WARNING: Finish Data COMFLICT at TP{0:03}".format(row[0].value))

def WriteSN():
    for row in wb["SN"].iter_rows(min_row=2, max_col=5, max_row=516):
        where = row[0].value-1
        c=1
        for n in CleanRoom[where].ESD_SN:
            row[c].value = n
            c=c+1


def Save():
    name = datetime.date.today().strftime("%Y_%m_%d")+".xlsx"
    wb.save(name)


if __name__ == "__main__":
    ReadData()
